#!/usr/bin/env python3
"""
生成带类型元数据+标记+注释行的测试用 Excel 配置文件。
Row 1: 表头（字段名）
Row 2: 类型（字段类型元数据）
Row 3: 元数据标记（C/S/CS, 范围, 默认值等）
Row 4: 注释说明
Row 5+: 数据
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "configs")

HEADERS = ["id", "name", "type", "quality", "max_stack", "use_type", "description", "notes"]
TYPES   = ["int", "string", "int", "int",    "int",      "int",     "string",    "string"]
METAS   = ["CS", "CS",     "CS",   "CS",     "CS[1,]",   "CS[0,2]","CS",        "X"]
COMMENTS= ["唯一ID", "物品名称", "类型ID", "品质等级(1-5)", "最大堆叠", "使用类型(0=无 1=消耗)", "描述", "策划备注(不导出)"]
DATA    = [
    [1001, "金币",    1, 1, 999999, 0, "游戏通用货币", "基础货币"],
    [1002, "钻石",    1, 3, 999999, 0, "稀有货币", ""],
    [2001, "生命药水", 2, 2, 99,     1, "恢复50点生命值", ""],
    [2002, "魔法药水", 2, 2, 99,     1, "恢复30点魔法值", ""],
    [3001, "铁剑",    3, 2, 1,      0, "一把普通的铁剑", "新手装备"],
]


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    type_font   = Font(italic=True, color="666666")
    type_fill   = PatternFill("solid", fgColor="D9E2F3")
    meta_font   = Font(color="996633")
    meta_fill   = PatternFill("solid", fgColor="FFF2CC")
    comment_font = Font(color="666666")
    comment_fill = PatternFill("solid", fgColor="F2F2F2")

    # Row 1: 表头
    for ci, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Row 2: 类型元数据
    for ci, t in enumerate(TYPES, start=1):
        cell = ws.cell(row=2, column=ci, value=t)
        cell.font = type_font
        cell.fill = type_fill

    # Row 3: 元数据标记
    for ci, m in enumerate(METAS, start=1):
        cell = ws.cell(row=3, column=ci, value=m)
        cell.font = meta_font
        cell.fill = meta_fill

    # Row 4: 注释
    for ci, c in enumerate(COMMENTS, start=1):
        cell = ws.cell(row=4, column=ci, value=c)
        cell.font = comment_font
        cell.fill = comment_fill

    # Row 5+: 数据
    for ri, row in enumerate(DATA, start=5):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)

    # 列宽
    for ci in range(1, len(HEADERS) + 1):
        col_letter = chr(64 + ci)
        max_len = max(
            len(str(HEADERS[ci-1])),
            len(str(COMMENTS[ci-1])),
        )
        for row in DATA:
            if row[ci-1] is not None:
                max_len = max(max_len, len(str(row[ci-1])))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    filepath = os.path.join(OUTPUT_DIR, "ItemConfig.xlsx")
    wb.save(filepath)
    wb.close()
    print(f"[OK] 生成 {filepath}")
    print(f"     表头: {HEADERS}")
    print(f"     类型: {TYPES}")
    print(f"     标记: {METAS}")
    print(f"     注释: {COMMENTS}")
    print(f"     数据: {len(DATA)} 行")


if __name__ == "__main__":
    main()
